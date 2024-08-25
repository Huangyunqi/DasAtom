from naqst import *
from openpyxl import Workbook
import math
import time
from vfsexp import Vf
from Enola.codegen import CodeGen, global_dict
from Enola.route import QuantumRouter
import json
from SA import find_map_SA

if __name__ == "__main__":

	'''path = "results/tetris/3_regular_graph/map_tetris/"
	files = os.listdir(path)
	wb = Workbook()
	ws = wb.active
	ws.append(['circuit name', 'Fidelity', 'inserted SWAP', 'gate_cycle'])
	for num_file in range(len(files)):
		file_name = files[num_file]
		if not file_name.endswith('.qasm'):
			continue
		print(file_name)
		cycle_file = file_name+'.txt'
		print(cycle_file)
		Fidelity, swap_count, gate_cycle = compute_fidelity_tetris(cycle_file, file_name, path)
		ws.append([file_name, Fidelity, swap_count, gate_cycle])

	wb.save(path+'quantum_volume_total.xlsx')'''

	#qasm input
	path_type = 'qft_cz'
	path = "Data/{}/circuits/".format(path_type)
	path_embeddings = "Data/{}/Rb2Re4/embeddings/".format(path_type)
	path_partitions = "Data/{}/Rb2Re4/partitions/".format(path_type)
	path_result = "results/yq_test/{}/Rb2Re4/".format(path_type)
	files = os.listdir(path)
	save_file_sub = True
	save_file_tot = True
	from_read = False
	write_txt = True
	#file_name = 'qft_50.qasm'
	total_wb = Workbook()
	total_ws = total_wb.active
	total_ws.append(['file name','Qubits','CZ_gates', 'depth', 'fidelity', 'movement_fidelity', 'movement times', 'gate cycles', 'partitions', 'Times'])
	for num_file in range(len(files)):
	#for num_file in [0]:
		#file_name = 'cz_2q_dj_indep_qiskit_{}.qasm'.format(num_file+5)
		file_name = 'cz_2q_qft_{}.qasm'.format(num_file+5)
		#file_name = files[num_file]
		#file_name = 'qft_cz_all.qasm'
		print(file_name)
		wb = Workbook()
		ws = wb.active
		log = []
		total_time = time.time()
		cz_circuit = CreateCircuitFromQASM(file_name, path)
	#transform to cz-based circuit
		#cz_circuit = transpile(circuit, basis_gates=['cz', 'rx', 'ry', 'rz', 'h', 't'])
	#cz gates list
		gate_2q_list = get_2q_gates_list(cz_circuit)
	#print(gate_2q_list)
	#obtain corresponding DAG
		cirr, dag = gates_list_to_QC(gate_2q_list)
		gate_num = len(gate_2q_list)
	#obtain the qubits number
		num_q = qubits_num(gate_2q_list)
		print("Num of gates", gate_num)
		log.append(['Num of gate', gate_num])
		arch_size = math.ceil(math.sqrt(num_q))
		#Rb = math.sqrt(2)
		log.append(['arch_size', 'sqrt(num_q)', arch_size])
		Rb = 2
		log.append(['Rb', Rb])
		r_re = 2*Rb
		log.append(['r_re', r_re])
	#obtain the corresponding coupling_graph 
		coupling_graph = generate_grid_with_Rb(arch_size,arch_size, Rb)

	#obtain the gates partition
		time_part = time.time()
		#ini_map = qasm_to_map('results/initial_map/'+file_name)
		#partition_gates = partition_from_ini(dag, coupling_graph, ini_map)
		if from_read:
			partition_gates = read_data(path_partitions, file_name.removesuffix(".qasm")+'.txt')
		else:
			partition_gates = parition_from_DAG(dag, coupling_graph)
			if write_txt:
				write_data(partition_gates, path_partitions, file_name.removesuffix(".qasm")+'part.txt')		
			time_part1 = time.time()
			print("partition time is, ",time_part1-time_part)
			log.append(["partition time", time_part1-time_part])

		part_gate_num = 0
		for gates in partition_gates:
	#		print("------")
			part_gate_num += len(gates)
	#	print(gates)
		print("final num is:", part_gate_num)


    #for each partition, find a proper embedding
		time_embed = time.time()

		if from_read:
			embeddings = read_data(path_embeddings, file_name.removesuffix(".qasm")+'.txt')
		else:
			embeddings, extend_pos = get_embeddings(partition_gates, coupling_graph, num_q, arch_size, Rb)
			if write_txt:
				write_data(embeddings, path_embeddings, file_name.removesuffix(".qasm")+'emb.txt')
			time_embed1 = time.time()
			print("partition number:", len(partition_gates))
			log.append(["find embeddings time", time_embed1-time_embed])
			if len(extend_pos) != 0:
				log.append(["extend graph times", len(extend_pos)])
				log.append(extend_pos)
				arch_size += len(extend_pos)

		parallel_gates = []
		time_paral = time.time()
		for i in range(len(partition_gates)):
			gates = get_parallel_gates(partition_gates[i], coupling_graph, embeddings[i], r_re)
			parallel_gates.append(gates)
		time_paral1 = time.time()
		log.append(["find parallel_gates time", time_paral1-time_paral])


		route = QuantumRouter(num_q, embeddings, partition_gates, [arch_size, arch_size])
		route.run()
		total_paralled = []
		all_movements = []
		for num in range(len(embeddings) - 1):
			log.append([str(embeddings[num])])
			for gates in parallel_gates[num]:
				log.append([str(gates[it]) for it in range(len(gates))])
				total_paralled.append(gates)
			for paral_moves in route.movement_list[num]:
				log.append([str(paral_moves[it]) for it in range(len(paral_moves))])
				all_movements.append(paral_moves)
		
		if len(partition_gates) > 1:
			log.append([str(embeddings[num+1])])
			for gates in parallel_gates[num+1]:
				log.append([str(gates[it]) for it in range(len(gates))])
				total_paralled.append(gates)
		else:
			log.append([str(embeddings[0])])
			for gates in parallel_gates[0]:
				log.append([str(gates[it]) for it in range(len(gates))])
				total_paralled.append(gates)

		t_idle, Fidelity, move_fidelity = compute_fidelity(total_paralled, all_movements, num_q, gate_num)
		print("Fidelity is:", Fidelity)
		log.append(["Fidelity:", Fidelity])
		log.append(["t_idle:", t_idle])
		log.append(["move_fidelity", move_fidelity])
		log.append(["Movement times", len(all_movements)])
		log.append(["parallel times", len(total_paralled)])
		log.append(["partitions", len(embeddings)])
		total_time1 = time.time()
		log.append(["total time:", total_time1-total_time])
		para = set_parameters(True)
		log_para = []
		for key, value in para.items():
			log_para.append(str(key))
			log_para.append(str(value))
		log.append(log_para)
		for item in log:
			#print(item)
			ws.append(item)

		total_ws.append([file_name, num_q, gate_num, cirr.depth(), Fidelity, move_fidelity, len(all_movements), len(total_paralled), len(embeddings), total_time1-total_time])
		save_file_name = path_result+'{}_rb{}_archsize{}_mini_dis.xlsx'.format(file_name, Rb, arch_size)
		print(save_file_name)
		if save_file_sub:
			wb.save(save_file_name)

	para = set_parameters(True)
	log_para = []
	for key, value in para.items():
		log_para.append(str(key))
		log_para.append(str(value))
	total_ws.append(log_para)
	if save_file_tot:
		total_wb.save(path_result+'{}_total.xlsx'.format(path_type))'''
	

	 	